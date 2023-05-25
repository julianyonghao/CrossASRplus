# Import module
import openpyxl
import difflib
import xlsxwriter

def post_processing(file_path, output_file_path):
    """
    To write difference in transcriptions ASRs produce from TTS-generated speech and human speech to an excel file. 
    """
    
    # Load excel with its path
    wrkbk = openpyxl.load_workbook(file_path)
    sh = wrkbk.active
    prev_filename = ''
    k = 1
    dict = {}
    row = 1
  
    # Iterate through excel and display data
    # Row 
    for i in range(2, sh.max_row+1):
        # Get filename
        filename_obj = sh.cell(row=i, column=1)
        filename = filename_obj.value

        if filename == prev_filename:
            prev_filename = filename 
            continue 
        else:
        
            # Initiate the Differ object
            d = difflib.Differ()

            # Calculate the difference between the two texts
            original = sh.cell(row=i, column=3).value
            transcribed = sh.cell(row=i, column=4).value
            diff = d.compare(original.split(), transcribed.split())

            # Output the result
            for word in list(diff):
                if word[0] == '-' or word[0] == '+':
                    print(word)
                    if word in dict:
                        dict[word] = dict[word] + 1
                    else:
                        dict[word] = 1
            prev_filename = filename

    print("Differences in words written to file: ", output_file_path)
    workbook = xlsxwriter.Workbook(output_file_path)
    worksheet = workbook.add_worksheet()

    # Write word and frequency
    worksheet.write(0,0,"word")
    worksheet.write(0,1, "frequency")

    for key, value in dict.items():
        worksheet.write(row,0,key)
        worksheet.write(row,1, value)
        row +=1
    workbook.close()
    
if __name__ == "__main__":
    post_processing("results/false_alarms.xlsx", "false_alarm_pattern.xlsx")